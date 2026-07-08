"""
Office document <-> PDF conversion.

Word/Excel/PowerPoint -> PDF requires an actual layout/rendering engine;
there is no pure-Python way to do this faithfully. We shell out to
LibreOffice in headless mode (`soffice --headless --convert-to pdf`), which
is free, open-source, and the industry-standard approach used by most
self-hosted document platforms (it's the same engine Google/Nextcloud/
Collabora rely on). No paid API, no cloud rendering service involved.

PDF -> Word uses `pdf2docx` (pure Python + PyMuPDF under the hood, MIT
licensed) which gives a much better editable-Word result than routing
through LibreOffice for that direction.
"""
import subprocess
from pathlib import Path

from app.config.settings import get_settings
from app.services.pdf.exceptions import PDFProcessingError

settings = get_settings()


def _run_soffice_convert(input_path: Path, output_dir: Path, target_format: str) -> Path:
    """Runs LibreOffice headless conversion. Raises PDFProcessingError with a clear message on failure."""
    cmd = [
        settings.SOFFICE_BIN,
        "--headless",
        "--norestore",
        "--convert-to", target_format,
        "--outdir", str(output_dir),
        str(input_path),
    ]
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=settings.SOFFICE_TIMEOUT_SECONDS,
        )
    except FileNotFoundError as exc:
        raise PDFProcessingError(
            "Document conversion engine (LibreOffice) is not installed on this server. "
            "See docs/DEPLOYMENT.md for setup instructions."
        ) from exc
    except subprocess.TimeoutExpired as exc:
        raise PDFProcessingError("Conversion timed out. The file may be too large or complex.") from exc

    if result.returncode != 0:
        raise PDFProcessingError(f"Conversion failed: {result.stderr.strip()[:300] or 'unknown error'}")

    expected_output = output_dir / f"{input_path.stem}.{target_format}"
    if not expected_output.exists():
        raise PDFProcessingError("Conversion completed but no output file was produced.")
    return expected_output


def word_to_pdf(input_path: Path, output_dir: Path) -> Path:
    return _run_soffice_convert(input_path, output_dir, "pdf")


def excel_to_pdf(input_path: Path, output_dir: Path) -> Path:
    return _run_soffice_convert(input_path, output_dir, "pdf")


def ppt_to_pdf(input_path: Path, output_dir: Path) -> Path:
    return _run_soffice_convert(input_path, output_dir, "pdf")


def pdf_to_excel(input_path: Path, output_dir: Path) -> Path:
    """Best-effort table extraction: pulls tabular text via PyMuPDF and writes an .xlsx via openpyxl."""
    import fitz
    from openpyxl import Workbook

    try:
        doc = fitz.open(input_path)
    except Exception as exc:
        raise PDFProcessingError("Could not read the PDF.") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    row_idx = 1
    try:
        for page in doc:
            tables = page.find_tables()
            if tables and tables.tables:
                for table in tables.tables:
                    for row in table.extract():
                        for col_idx, value in enumerate(row, start=1):
                            ws.cell(row=row_idx, column=col_idx, value=value)
                        row_idx += 1
                    row_idx += 1
            else:
                text = page.get_text().strip()
                if text:
                    ws.cell(row=row_idx, column=1, value=text)
                    row_idx += 1
    finally:
        doc.close()

    output_path = output_dir / f"{input_path.stem}.xlsx"
    wb.save(output_path)
    return output_path


def pdf_to_word(input_path: Path, output_dir: Path) -> Path:
    from pdf2docx import Converter

    output_path = output_dir / f"{input_path.stem}.docx"
    cv = Converter(str(input_path))
    try:
        cv.convert(str(output_path))
    except Exception as exc:
        raise PDFProcessingError("Could not convert this PDF to Word. It may be scanned/image-only or password protected.") from exc
    finally:
        cv.close()
    return output_path
